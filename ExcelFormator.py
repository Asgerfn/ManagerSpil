import pandas as pd
from openpyxl import Workbook, load_workbook
from HelpFunctions import *
from WipeExcel import AntalSpillere


wb = load_workbook("kamp.xlsx")
ws = wb.active 

def listOfPeople(): 
    wb = load_workbook("kamp.xlsx")
    ws = wb.active 
    listOfNames = []
    for i in range(1,AntalSpillere+1): 
        listOfNames.append(ws['A'+str(i)].value)
    return(listOfNames)
Liste = listOfPeople()

def people(navn,liste): 
    for i in range(len(liste)): 
        if navn == liste[i]:  
            return(i)

def NavneLængde(liste): 
    List = []
    for i in range(len(liste)): 
        List.append(len(liste[i])) 
    return(List)
NavneLængdeListe = (NavneLængde(Liste))


def updateBoard(i:int,value:int):  
    wb = load_workbook("kamp.xlsx")
    ws = wb.active 
    ws['B'+str(i+1)].value = int(ws['B'+str(i+1)].value)+value
    wb.save("kamp.xlsx")


def længdeExcel(): 
    wb = load_workbook("managerspil.xlsx")
    ws = wb.active
    counter = 0
    for i in range(1,1000): 
        if ws['A'+str(i)].value == None: 
            counter= counter+1
            if counter == 20: 
                return(i) 
        else: 
            counter = 0
        
længdeexcel = (længdeExcel())
def newExcel(x): 
    Liste = listOfPeople()
    wbb = load_workbook("managerspil.xlsx")
    wb = load_workbook("kamp.xlsx")
    ws = wb.active
    wa = wbb.active
    for i in range(1,x): 
        for j in range(len(Liste)):
            if wa['A'+str(i)].value == Liste[j]: 
                for A in range(1,AntalSpillere+1): 
                    if ws['A'+str(A)].value == Liste[j]: 
                        wa['B'+str(i)].value = ws['B'+str(A)].value
    wbb.save('managerspil.xlsx')
    
    
    

print(newExcel(længdeexcel))
