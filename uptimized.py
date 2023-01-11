
from tokenize import Ignore
from selenium import webdriver
import re
from selenium.webdriver.common.keys import Keys
import openpyxl
import numpy as np
from openpyxl import Workbook, load_workbook
from ExcelFormator import *
from HelpFunctions import *
from WipeExcel import *

HjemmmeBaneLink = ["https://www.badmintonplayer.dk/DBF/HoldTurnering/Stilling/#5,2022,14816,1,1,,425312,","https://www.badmintonplayer.dk/DBF/HoldTurnering/Stilling/#5,2022,14816,1,1,,425322,","https://www.badmintonplayer.dk/DBF/HoldTurnering/Stilling/#5,2022,14816,1,1,,425324,"]
UdebaneLink = ["https://www.badmintonplayer.dk/DBF/HoldTurnering/Stilling/#5,2022,14816,1,1,,425319,","https://www.badmintonplayer.dk/DBF/HoldTurnering/Stilling/#5,2022,14816,1,1,,425329,","https://www.badmintonplayer.dk/DBF/HoldTurnering/Stilling/#5,2022,14829,1,1,,425533,"]





def udebane(x,tekst): 
    wb = load_workbook("kamp.xlsx") 
    ws = wb.active
    Liste = listOfPeople()
    NavneLængdeListe = (NavneLængde(Liste))
    brugteNavne = []
    for i in range(500,len(x)-13): 
        for j in range(1,len(Liste)): 
            if (x[i:i+NavneLængdeListe[j]] == Liste[j]) and (Liste[j] not in brugteNavne):
                excelNummer = people(Liste[j],Liste)+1 
                brugteNavne.append(Liste[j])
                B,c,d,e,f,Exceli,Excelj,l,m,o,p,r,s,t,u,Excelx,y,ad,ae,af,ag,ah,ai,ak,al,an,ao,aq,ar,at,au,av,aw=(helpUdebane(x,Liste[j],excelNummer,tekst))
                ws['C'+str(excelNummer)].value = ws['C'+str(excelNummer)].value+c
                ws['D'+str(excelNummer)].value = ws['D'+str(excelNummer)].value+d
                ws['B'+str(excelNummer)].value = ws['B'+str(excelNummer)].value+B
                ws['E'+str(excelNummer)].value = ws['E'+str(excelNummer)].value+e
                ws['F'+str(excelNummer)].value = ws['F'+str(excelNummer)].value+f
                ws['I'+str(excelNummer)].value = ws['I'+str(excelNummer)].value+Exceli
                ws['J'+str(excelNummer)].value = ws['J'+str(excelNummer)].value+Excelj
                ws['L'+str(excelNummer)].value = ws['L'+str(excelNummer)].value+l
                ws['M'+str(excelNummer)].value = ws['M'+str(excelNummer)].value+m
                ws['O'+str(excelNummer)].value = ws['O'+str(excelNummer)].value+o
                ws['P'+str(excelNummer)].value = ws['P'+str(excelNummer)].value+p
                ws['R'+str(excelNummer)].value = ws['R'+str(excelNummer)].value+r
                ws['S'+str(excelNummer)].value = ws['S'+str(excelNummer)].value+s
                ws['T'+str(excelNummer)].value = ws['T'+str(excelNummer)].value+t
                ws['U'+str(excelNummer)].value = ws['U'+str(excelNummer)].value+u
                ws['X'+str(excelNummer)].value = ws['X'+str(excelNummer)].value+Excelx
                ws['Y'+str(excelNummer)].value = ws['Y'+str(excelNummer)].value+y
                ws['AD'+str(excelNummer)].value = ws['AD'+str(excelNummer)].value+ad
                ws['AE'+str(excelNummer)].value = ws['AE'+str(excelNummer)].value+ae
                ws['AF'+str(excelNummer)].value = ws['AF'+str(excelNummer)].value+af
                ws['AG'+str(excelNummer)].value = ws['AG'+str(excelNummer)].value+ag
                ws['AH'+str(excelNummer)].value = ws['AH'+str(excelNummer)].value+ah
                ws['AI'+str(excelNummer)].value = ws['AI'+str(excelNummer)].value+ai
                ws['AK'+str(excelNummer)].value = ws['AK'+str(excelNummer)].value+ak
                ws['AL'+str(excelNummer)].value = ws['AL'+str(excelNummer)].value+al
                ws['AN'+str(excelNummer)].value = ws['AN'+str(excelNummer)].value+an
                ws['AO'+str(excelNummer)].value = ws['AO'+str(excelNummer)].value+ao
                ws['AQ'+str(excelNummer)].value = ws['AQ'+str(excelNummer)].value+aq
                ws['AR'+str(excelNummer)].value = ws['AR'+str(excelNummer)].value+ar
                ws['AT'+str(excelNummer)].value = ws['AT'+str(excelNummer)].value+at
                ws['AU'+str(excelNummer)].value = ws['AU'+str(excelNummer)].value+au
                ws['AV'+str(excelNummer)].value = ws['AV'+str(excelNummer)].value+av
                ws['AW'+str(excelNummer)].value = ws['AW'+str(excelNummer)].value+aw
                break
    wb.save("kamp.xlsx")   
    newExcel(længdeexcel)
    return()
#print(udebane(finalTekst))

def Hjemmebane(x,tekst): 
    wb = load_workbook("kamp.xlsx")
    ws = wb.active
    Liste = listOfPeople()
    NavneLængdeListe = (NavneLængde(Liste))
    brugteNavne = []

    for i in range(500,len(x)-13): 
        for j in range(1,len(Liste)): 
            
            if (x[i:i+NavneLængdeListe[j]] == Liste[j]) and (Liste[j] not in brugteNavne):
                excelNummer = people(Liste[j],Liste)+1 
                B,c,d,e,f,Exceli,Excelj,l,m,o,p,r,s,t,u,Excelx,y,ad,ae,af,ag,ah,ai,ak,al,an,ao,aq,ar,at,au,av,aw=(helpHjemmebane(x,Liste[j],excelNummer,tekst))
                ws['C'+str(excelNummer)].value = ws['C'+str(excelNummer)].value+c
                ws['D'+str(excelNummer)].value = ws['D'+str(excelNummer)].value+d
                ws['B'+str(excelNummer)].value = ws['B'+str(excelNummer)].value+B
                ws['E'+str(excelNummer)].value = ws['E'+str(excelNummer)].value+e
                ws['F'+str(excelNummer)].value = ws['F'+str(excelNummer)].value+f
                ws['I'+str(excelNummer)].value = ws['I'+str(excelNummer)].value+Exceli
                ws['J'+str(excelNummer)].value = ws['J'+str(excelNummer)].value+Excelj
                ws['L'+str(excelNummer)].value = ws['L'+str(excelNummer)].value+l
                ws['M'+str(excelNummer)].value = ws['M'+str(excelNummer)].value+m
                ws['O'+str(excelNummer)].value = ws['O'+str(excelNummer)].value+o
                ws['P'+str(excelNummer)].value = ws['P'+str(excelNummer)].value+p
                ws['R'+str(excelNummer)].value = ws['R'+str(excelNummer)].value+r
                ws['S'+str(excelNummer)].value = ws['S'+str(excelNummer)].value+s
                ws['T'+str(excelNummer)].value = ws['T'+str(excelNummer)].value+t
                ws['U'+str(excelNummer)].value = ws['U'+str(excelNummer)].value+u
                ws['X'+str(excelNummer)].value = ws['X'+str(excelNummer)].value+Excelx
                ws['Y'+str(excelNummer)].value = ws['Y'+str(excelNummer)].value+y
                ws['AD'+str(excelNummer)].value = ws['AD'+str(excelNummer)].value+ad
                ws['AE'+str(excelNummer)].value = ws['AE'+str(excelNummer)].value+ae
                ws['AF'+str(excelNummer)].value = ws['AF'+str(excelNummer)].value+af
                ws['AG'+str(excelNummer)].value = ws['AG'+str(excelNummer)].value+ag
                ws['AH'+str(excelNummer)].value = ws['AH'+str(excelNummer)].value+ah
                ws['AI'+str(excelNummer)].value = ws['AI'+str(excelNummer)].value+ai
                ws['AK'+str(excelNummer)].value = ws['AK'+str(excelNummer)].value+ak
                ws['AL'+str(excelNummer)].value = ws['AL'+str(excelNummer)].value+al
                ws['AN'+str(excelNummer)].value = ws['AN'+str(excelNummer)].value+an
                ws['AO'+str(excelNummer)].value = ws['AO'+str(excelNummer)].value+ao
                ws['AR'+str(excelNummer)].value = ws['AR'+str(excelNummer)].value+ar
                ws['AQ'+str(excelNummer)].value = ws['AQ'+str(excelNummer)].value+aq
                ws['AT'+str(excelNummer)].value = ws['AT'+str(excelNummer)].value+at
                ws['AU'+str(excelNummer)].value = ws['AU'+str(excelNummer)].value+au
                ws['AV'+str(excelNummer)].value = ws['AV'+str(excelNummer)].value+av
                ws['AW'+str(excelNummer)].value = ws['AW'+str(excelNummer)].value+aw
                brugteNavne.append(Liste[j])
                break
    wb.save("kamp.xlsx")  
    newExcel(længdeexcel) 
    return()
#print(Hjemmebane(finalTekst))

 
Hjemmelinjer = []
Hjemmefil = open("Hjemmebane.txt", "r")
Hjemmelinjer = Hjemmefil.readlines()
gentagHjemmeliste = []
for line in Hjemmelinjer:
    gentagHjemmeliste.append((line.strip()))
    
    
with open("Hjemmebane.txt", "a") as Hjemmebanen: 
    for i in range(len(HjemmmeBaneLink)): 
        if HjemmmeBaneLink[i] not in gentagHjemmeliste:
            Hjemmebanen.write(HjemmmeBaneLink[i] +"\n") 

  

Udelinjer = []
udefil = open("Udebane.txt", "r")
Udelinjer = udefil.readlines()
gentagUdeliste = []
for line in Udelinjer:
    gentagUdeliste.append((line.strip()))

with open("Udebane.txt", "a") as Udebanen: 
    for i in range(len(UdebaneLink)): 
        if UdebaneLink[i] not in gentagUdeliste:
            Udebanen.write(UdebaneLink[i] +"\n")


brugtelinks = open("BrugteLinks.txt","a")
linjebrugt = open("BrugteLinks.txt","r")
Brugtelinjer = linjebrugt.readlines()



def kørSpil(x,y,link,brugtLinks): 
    for i in range(len(x)): 
        if x[i] not in brugtLinks: 
                
                print("Hjemmebane holdkamp kører... \n")
                print("Nummer ", i,"ud af ", len(x))
                PATH = "C:\webdrivers\chromedriver.exe" 
                driver = webdriver.Chrome(PATH)
                driver.get(x[i])
                main = driver.find_element_by_id("aspnetForm")
                tekst = str(main.text)
                finalTekst = (fjernNavne(tekst))
                Hjemmebane(finalTekst,tekst)
                
                print("Hejmmebane holdkamp opdateret nummer",i, "\n")
                if x[i] not in brugtLinks: 
                    link.write(x[i])
    for j in range(len(y)): 
            if y[j] not in brugtLinks: 
                
                print("Udebane holdkamp kører... \n")
                print("Nummer ", j,"ud af ", len(y))
                PATH = "C:\webdrivers\chromedriver.exe" 
                driver = webdriver.Chrome(PATH)
                driver.get(y[j])
                main = driver.find_element_by_id("aspnetForm")
                tekst = str(main.text)
                finalTekst = (fjernNavne(tekst))
                
                udebane(finalTekst,tekst)
                print("Udebane holdkamp opdateret \n")
                if y[j] not in brugtLinks:
                    link.write(y[j])
    return("så er spillet klar!")
print(kørSpil(Hjemmelinjer,Udelinjer,brugtelinks,Brugtelinjer))

