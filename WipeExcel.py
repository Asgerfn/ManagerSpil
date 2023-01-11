import pandas as pd
import os
import openpyxl
from openpyxl import Workbook, load_workbook


def LengthOfExcel():
    wb = load_workbook("kamp.xlsx")
    ws = wb.active 
    AntalSpillere = 0
    for i in range(2,500): 
        val = ws['A'+str(i)].value
        if val is not None:
            AntalSpillere += 1
    return(AntalSpillere)

AntalSpillere = (LengthOfExcel())+1

def slette(x):
    wb = load_workbook("kamp.xlsx")
    ws = wb.active 
    for i in range(2,AntalSpillere+2): 
        ws['B'+str(i)].value = 0
        ws['C'+str(i)].value = 0
        ws['D'+str(i)].value = 0
        ws['E'+str(i)].value = 0
        ws['F'+str(i)].value = 0
        ws['I'+str(i)].value = 0
        ws['J'+str(i)].value = 0
        ws['L'+str(i)].value = 0
        ws['M'+str(i)].value = 0
        ws['O'+str(i)].value = 0
        ws['P'+str(i)].value = 0
        ws['R'+str(i)].value = 0
        ws['S'+str(i)].value = 0
        ws['T'+str(i)].value = 0
        ws['U'+str(i)].value = 0
        ws['X'+str(i)].value = 0
        ws['Y'+str(i)].value = 0
        ws['AD'+str(i)].value = 0
        ws['AE'+str(i)].value = 0
        ws['AF'+str(i)].value = 0
        ws['AG'+str(i)].value = 0
        ws['AH'+str(i)].value = 0
        ws['AI'+str(i)].value = 0
        ws['AL'+str(i)].value = 0
        ws['AK'+str(i)].value = 0
        ws['AN'+str(i)].value = 0
        ws['AO'+str(i)].value = 0
        ws['AQ'+str(i)].value = 0
        ws['AR'+str(i)].value = 0
        ws['AU'+str(i)].value = 0
        ws['AT'+str(i)].value = 0
        ws['AV'+str(i)].value = 0
        ws['AW'+str(i)].value = 0
    open("BrugteLinks.txt", 'w').close()
    
    print("slet")
    wb.save("kamp.xlsx")


print(slette(1))